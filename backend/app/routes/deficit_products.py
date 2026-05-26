from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlmodel import Session

from app.db import get_session
from app.models import User
from app.security import get_current_user
from app.services.product_intelligence.parsers.uzum_parser import parse_uzum
from app.services.product_intelligence.uzum_analyzer import analyze_items

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


router = APIRouter(tags=["deficit-products"])

REPORTS_DIR = Path("/var/www/marketcard/generated_reports")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

COST_BY_LIMIT = {
    100: 20,
    300: 20,
    500: 20,
}


class DeficitProductsRequest(BaseModel):
    query: str
    limit: int = 100
    marketplace: str = "uzum"


def _num(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(float(value))
    except Exception:
        return default


def _rating(value: Any) -> float:
    try:
        if value is None:
            return 0.0
        return round(float(value), 1)
    except Exception:
        return 0.0


def _score_item(item: dict[str, Any], stats: dict[str, Any]) -> dict[str, Any]:
    price = _num(item.get("price"))
    reviews = _num(item.get("reviews"))
    rating = _rating(item.get("rating"))

    avg_price = _num(stats.get("avg_price"))
    min_price = _num(stats.get("min_price"))
    max_price = _num(stats.get("max_price"))
    competitors = _num(stats.get("count"))
    sellers = _num(stats.get("sellers"))

    demand_points = 0
    if reviews >= 500:
        demand_points = 35
    elif reviews >= 150:
        demand_points = 28
    elif reviews >= 50:
        demand_points = 20
    elif reviews >= 10:
        demand_points = 12
    else:
        demand_points = 6

    rating_points = 15 if rating >= 4.7 else 11 if rating >= 4.4 else 7 if rating >= 4.0 else 3

    competition_penalty = 0
    if competitors >= 80:
        competition_penalty = 25
    elif competitors >= 50:
        competition_penalty = 18
    elif competitors >= 25:
        competition_penalty = 10
    else:
        competition_penalty = 3

    price_gap_points = 0
    if avg_price and price:
        if price <= avg_price * 0.9:
            price_gap_points = 18
        elif price <= avg_price:
            price_gap_points = 12
        elif price <= avg_price * 1.15:
            price_gap_points = 7
        else:
            price_gap_points = 2

    deficit_points = 0
    if competitors <= 25 and reviews >= 50:
        deficit_points = 25
    elif competitors <= 50 and reviews >= 150:
        deficit_points = 18
    elif competitors <= 80 and reviews >= 300:
        deficit_points = 10

    score = demand_points + rating_points + price_gap_points + deficit_points - competition_penalty
    score = max(0, min(100, score))

    if score >= 75:
        deficit = "Да"
        recommendation = "Дефицитный товар: можно тестировать, если сделать сильную карточку и цену около среднего рынка."
    elif score >= 60:
        deficit = "Возможно"
        recommendation = "Есть потенциал: проверьте поставщика, маржу и качество карточек конкурентов."
    else:
        deficit = "Нет"
        recommendation = "Рискованно: спрос/конкуренция не дают сильного преимущества."

    if competitors >= 80:
        competition = "высокая"
    elif competitors >= 30:
        competition = "средняя"
    else:
        competition = "низкая"

    if reviews >= 500:
        demand = "высокий"
    elif reviews >= 100:
        demand = "средний/высокий"
    elif reviews >= 20:
        demand = "средний"
    else:
        demand = "низкий"

    return {
        "product_id": str(item.get("product_id") or "").strip(),
        "title": str(item.get("title") or "").strip(),
        "niche": str(item.get("category") or "").strip(),
        "url": str(item.get("url") or "").strip(),
        "price": price,
        "avg_price": avg_price,
        "min_price": min_price,
        "max_price": max_price,
        "competitors": competitors,
        "sellers": sellers,
        "rating": rating,
        "reviews": reviews,
        "demand": demand,
        "competition": competition,
        "deficit": deficit,
        "score": score,
        "recommendation": recommendation,
    }


def _make_excel(rows: list[dict[str, Any]], query: str, limit: int) -> tuple[str, str]:
    now = datetime.now()
    safe_query = "".join(ch if ch.isalnum() else "_" for ch in query.lower())[:40].strip("_") or "uzum"
    filename = f"deficit_products_{safe_query}_{limit}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = REPORTS_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Дефицитные товары"

    headers = [
        "№",
        "Product ID",
        "Название товара",
        "Ниша / категория",
        "Ссылка на товар",
        "Текущая цена",
        "Средняя цена по нише",
        "Минимальная цена конкурентов",
        "Максимальная цена конкурентов",
        "Количество конкурентов",
        "Количество продавцов",
        "Рейтинг",
        "Отзывы",
        "Спрос",
        "Конкуренция",
        "Дефицит",
        "Opportunity Score",
        "Рекомендация",
        "Дата анализа",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(rows, start=1):
        ws.append([
            idx,
            row["product_id"],
            row["title"],
            row["niche"],
            row["url"],
            row["price"],
            row["avg_price"],
            row["min_price"],
            row["max_price"],
            row["competitors"],
            row["sellers"],
            row["rating"],
            row["reviews"],
            row["demand"],
            row["competition"],
            row["deficit"],
            row["score"],
            row["recommendation"],
            now.strftime("%Y-%m-%d %H:%M"),
        ])

        link_cell = ws.cell(row=idx + 1, column=5)
        if row["url"]:
            link_cell.hyperlink = row["url"]
            link_cell.value = "Открыть товар"
            link_cell.style = "Hyperlink"

    widths = {
        1: 8,
        2: 16,
        3: 45,
        4: 24,
        5: 20,
        6: 16,
        7: 20,
        8: 24,
        9: 24,
        10: 22,
        11: 20,
        12: 12,
        13: 12,
        14: 18,
        15: 18,
        16: 14,
        17: 18,
        18: 55,
        19: 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filename, f"/generated_reports/{filename}"


@router.post("/deficit-products")
def deficit_products(
    req: DeficitProductsRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    query = (req.query or "").strip()
    if len(query) < 2:
        raise HTTPException(status_code=400, detail="Введите категорию или поисковый запрос")

    limit = int(req.limit or 100)
    if limit not in COST_BY_LIMIT:
        raise HTTPException(status_code=400, detail="Доступные лимиты: 100, 300, 500")

    if (req.marketplace or "uzum").lower() != "uzum":
        raise HTTPException(status_code=400, detail="Сейчас доступен только Uzum")

    cost = COST_BY_LIMIT[limit]

    if (current_user.audit_credits or 0) < cost:
        raise HTTPException(
            status_code=402,
            detail=f"Недостаточно аудитов. Нужно {cost}, доступно {current_user.audit_credits or 0}",
        )

    items = parse_uzum(query, query, limit)

    if not items:
        raise HTTPException(
            status_code=502,
            detail="Uzum не вернул товары. Попробуйте другой запрос.",
        )

    stats = analyze_items(items)
    scored = [_score_item(item, stats) for item in items]

    deficit_rows = sorted(scored, key=lambda x: x["score"], reverse=True)[:limit]

    if not deficit_rows:
        raise HTTPException(
            status_code=404,
            detail="Товары не найдены по этому запросу.",
        )

    filename, download_url = _make_excel(deficit_rows, query, limit)

    current_user.audit_credits = max((current_user.audit_credits or 0) - cost, 0)
    session.add(current_user)
    session.commit()
    session.refresh(current_user)

    return {
        "success": True,
        "marketplace": "uzum",
        "query": query,
        "limit": limit,
        "cost": cost,
        "rows_found": len(deficit_rows),
        "download_url": download_url,
        "filename": filename,
        "audit_credits_left": current_user.audit_credits,
    }
